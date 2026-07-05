import json
import os
import sys
from datetime import datetime

from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle


NAVY = "17365D"
BLUE = "D9EAF7"
LIGHT = "F5F9FC"
GREEN = "D9EAD3"
AMBER = "FFF2CC"
RED = "F4CCCC"
GRAY = "E7E6E6"
TEXT = "1F2937"


def txt(value, fallback=""):
    if value is None:
        return fallback
    value = str(value).strip()
    return value if value else fallback


def num(value):
    try:
        return float(value or 0)
    except Exception:
        return 0


def whole(value):
    return int(round(num(value)))


def generated(value):
    try:
        return datetime.fromisoformat(str(value).replace("Z", "+00:00")).strftime("%Y-%m-%d %H:%M")
    except Exception:
        return txt(value)


def safe_money(value):
    if value in (None, ""):
        return ""
    return num(value)


def xlsx(data, out):
    agency = data.get("agency") or {}
    report = data.get("report") or {}
    charts = report.get("charts") or {}
    tables = report.get("tables") or {}
    employee = report.get("employeeSummary") or {}
    plantilla = report.get("plantillaSummary") or {}

    wb = Workbook()
    ws = wb.active
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="B7C9D6")

    ws.merge_cells("A1:L1")
    ws["A1"] = txt(agency.get("name"), "Southern Tagalog Regional Hospital")
    ws["A1"].font = Font(name="Arial", size=16, bold=True, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=NAVY)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:L2")
    ws["A2"] = "PERSONNEL STATISTICS & PLANTILLA ANALYTICS REPORT"
    ws["A2"].font = Font(name="Arial", size=13, bold=True, color=NAVY)
    ws["A2"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A3:L3")
    ws["A3"] = "Generated from current encoded HRIS records as of " + generated(report.get("generatedAt"))
    ws["A3"].font = Font(name="Arial", size=9, italic=True, color="666666")
    ws["A3"].alignment = Alignment(horizontal="center")

    cards = [
        ("Total Employees", whole(employee.get("total")), BLUE),
        ("Active Employees", whole(employee.get("active")), GREEN),
        ("Regular/Permanent", whole(employee.get("regular")), BLUE),
        ("JO/COS/Contract", whole(employee.get("nonPlantilla")), AMBER),
        ("Authorized Items", whole(plantilla.get("authorized")), BLUE),
        ("Active Items", whole(plantilla.get("active")), GREEN),
        ("Occupied Items", whole(plantilla.get("occupied")), GREEN),
        ("Vacant Items", whole(plantilla.get("vacant")), RED),
        ("Vacancy Rate", f"{num(plantilla.get('vacancyRate')):.1f}%", RED if num(plantilla.get("vacancyRate")) else GREEN),
    ]
    row = 5
    for i, (label, value, fill) in enumerate(cards):
        col = 1 + (i % 3) * 4
        row = 5 + (i // 3) * 3
        ws.merge_cells(start_row=row, start_column=col, end_row=row, end_column=col + 3)
        ws.cell(row, col, label)
        ws.cell(row, col).font = Font(name="Arial", size=9, bold=True, color=NAVY)
        ws.cell(row, col).fill = PatternFill("solid", fgColor=fill)
        ws.cell(row, col).alignment = Alignment(horizontal="center")
        ws.merge_cells(start_row=row + 1, start_column=col, end_row=row + 1, end_column=col + 3)
        ws.cell(row + 1, col, value)
        ws.cell(row + 1, col).font = Font(name="Arial", size=18, bold=True, color=TEXT)
        ws.cell(row + 1, col).alignment = Alignment(horizontal="center")
        for c in range(col, col + 4):
            ws.cell(row, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)
            ws.cell(row + 1, c).border = Border(top=thin, bottom=thin, left=thin, right=thin)

    def write_table(sheet, start_row, title, rows, headers):
        sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
        sheet.cell(start_row, 1, title)
        sheet.cell(start_row, 1).font = Font(name="Arial", size=12, bold=True, color=NAVY)
        header_row = start_row + 1
        for col, header in enumerate(headers, 1):
            cell = sheet.cell(header_row, col, header)
            cell.fill = PatternFill("solid", fgColor=NAVY)
            cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
            cell.border = Border(bottom=thin)
        for r_idx, row_data in enumerate(rows, header_row + 1):
            for col, key in enumerate(headers, 1):
                source_key = key[0].lower() + key[1:].replace(" ", "")
                value = row_data.get(source_key, row_data.get(key, ""))
                cell = sheet.cell(r_idx, col, value)
                cell.font = Font(name="Arial", size=9)
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(bottom=Side(style="hair", color="D9D9D9"))
            if r_idx % 2 == 0:
                for col in range(1, len(headers) + 1):
                    sheet.cell(r_idx, col).fill = PatternFill("solid", fgColor=LIGHT)
        return header_row + max(1, len(rows)) + 2

    detail = wb.create_sheet("Analytics")
    detail.sheet_view.showGridLines = False
    r = 1
    r = write_simple_series(detail, r, "Employees by Department", charts.get("byDepartment", []), ["Label", "Active", "Inactive", "Total"])
    r = write_simple_series(detail, r, "Employees by Employment Status", charts.get("byEmploymentStatus", []), ["Label", "Active", "Inactive", "Total"])
    r = write_simple_series(detail, r, "Employees by Age Group", charts.get("byAgeGroup", []), ["Label", "Total"])
    r = write_simple_series(detail, r, "Plantilla by Division", charts.get("plantillaByDivision", []), ["Label", "Occupied", "Vacant", "Active", "Total"])
    r = write_simple_series(detail, r, "Plantilla by Salary Grade", charts.get("plantillaBySalaryGrade", []), ["Label", "Occupied", "Vacant", "Active", "Total"])

    make_bar_chart(ws, detail, "Employees by Department", 1, 11, 1, min(12, len(charts.get("byDepartment", []))), "A14")
    make_bar_chart(ws, detail, "Plantilla Occupancy by Division", 1, 11, 4, min(12, len(charts.get("plantillaByDivision", []))), "G14")

    items = wb.create_sheet("Plantilla Items")
    items.sheet_view.showGridLines = False
    headers = [
        "Item Number",
        "Position Title",
        "Salary Grade",
        "Salary Step",
        "Salary Amount",
        "Division",
        "Section",
        "Plantilla Type",
        "Item Status",
        "Occupancy Status",
        "Occupant Name",
        "Occupant No",
    ]
    items.append(headers)
    for cell in items[1]:
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row_data in tables.get("plantillaItems", []):
        items.append([
            row_data.get("itemNumber", ""),
            row_data.get("positionTitle", ""),
            row_data.get("salaryGrade", ""),
            row_data.get("salaryStep", ""),
            safe_money(row_data.get("salaryAmount")),
            row_data.get("division", ""),
            row_data.get("section", ""),
            row_data.get("plantillaType", ""),
            row_data.get("itemStatus", ""),
            row_data.get("occupancyStatus", ""),
            row_data.get("occupantName", ""),
            row_data.get("occupantNo", ""),
        ])
    for row_cells in items.iter_rows(min_row=2):
        for cell in row_cells:
            cell.font = Font(name="Arial", size=9)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(bottom=Side(style="hair", color="D9D9D9"))
        row_index = row_cells[0].row
        if row_index % 2 == 0:
            for cell in row_cells:
                cell.fill = PatternFill("solid", fgColor=LIGHT)
        if row_cells[9].value == "Vacant":
            row_cells[9].fill = PatternFill("solid", fgColor=RED)
    items.freeze_panes = "A2"
    items.auto_filter.ref = f"A1:L{max(1, items.max_row)}"

    for sheet in wb.worksheets:
        for col in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col)].width = min(35, max(11, max(len(txt(sheet.cell(row, col).value)) for row in range(1, min(sheet.max_row, 80) + 1)) + 2))
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        sheet.sheet_properties.pageSetUpPr.fitToPage = True

    wb.save(out)


def write_simple_series(sheet, start_row, title, rows, headers):
    sheet.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=len(headers))
    sheet.cell(start_row, 1, title)
    sheet.cell(start_row, 1).font = Font(name="Arial", size=12, bold=True, color=NAVY)
    header_row = start_row + 1
    for col, header in enumerate(headers, 1):
        cell = sheet.cell(header_row, col, header)
        cell.fill = PatternFill("solid", fgColor=NAVY)
        cell.font = Font(name="Arial", size=9, bold=True, color="FFFFFF")
    for index, row in enumerate(rows, header_row + 1):
        values = [
            row.get("label", ""),
            row.get("active", row.get("occupied", row.get("total", 0))),
        ]
        if len(headers) >= 3:
            values.append(row.get("inactive", row.get("vacant", 0)))
        if len(headers) >= 4:
            values.append(row.get("active", row.get("total", 0)))
        if len(headers) >= 5:
            values.append(row.get("total", 0))
        for col, value in enumerate(values[: len(headers)], 1):
            sheet.cell(index, col, value)
    return header_row + max(1, len(rows)) + 3


def make_bar_chart(target, source, title, min_col, max_col, start_row, row_count, anchor):
    if row_count <= 0:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = title
    chart.y_axis.title = ""
    chart.x_axis.title = "Count"
    data = Reference(source, min_col=2, max_col=min(max_col, source.max_column), min_row=start_row + 1, max_row=start_row + row_count + 1)
    cats = Reference(source, min_col=1, min_row=start_row + 2, max_row=start_row + row_count + 1)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 7
    chart.width = 14
    target.add_chart(chart, anchor)


def pdf(data, out):
    agency = data.get("agency") or {}
    report = data.get("report") or {}
    employee = report.get("employeeSummary") or {}
    plantilla = report.get("plantillaSummary") or {}
    charts = report.get("charts") or {}
    tables = report.get("tables") or {}

    doc = SimpleDocTemplate(
        out,
        pagesize=landscape(A4),
        leftMargin=10 * mm,
        rightMargin=10 * mm,
        topMargin=9 * mm,
        bottomMargin=10 * mm,
        title="Personnel Statistics and Plantilla Analytics",
    )
    styles = getSampleStyleSheet()
    title = ParagraphStyle("ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=15, leading=17, alignment=TA_CENTER, textColor=colors.HexColor("#17365D"))
    subtitle = ParagraphStyle("Subtitle", parent=styles["BodyText"], fontSize=7.5, leading=9, alignment=TA_CENTER, textColor=colors.HexColor("#666666"))
    section = ParagraphStyle("Section", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=10.5, leading=12, textColor=colors.HexColor("#17365D"))
    small = ParagraphStyle("Small", parent=styles["BodyText"], fontSize=6.5, leading=8)

    story = [
        Paragraph(txt(agency.get("name"), "Southern Tagalog Regional Hospital"), title),
        Paragraph("PERSONNEL STATISTICS & PLANTILLA ANALYTICS REPORT", title),
        Paragraph("Generated from current encoded HRIS records as of " + generated(report.get("generatedAt")), subtitle),
        Spacer(1, 4 * mm),
    ]

    cards = [
        ("Employees", whole(employee.get("total")), "Total encoded personnel"),
        ("Active", whole(employee.get("active")), "Active employee records"),
        ("Regular/Permanent", whole(employee.get("regular")), "Permanent or regular status"),
        ("Authorized Items", whole(plantilla.get("authorized")), "Plantilla item records"),
        ("Occupied", whole(plantilla.get("occupied")), "Active occupied items"),
        ("Vacant", whole(plantilla.get("vacant")), f"{num(plantilla.get('vacancyRate')):.1f}% vacancy rate"),
    ]
    card_rows = []
    for label, value, note in cards:
        card_rows.append([Paragraph(f"<b>{label}</b>", small), Paragraph(f"<font size='16'><b>{value}</b></font>", small), Paragraph(note, small)])
    card_table = Table(card_rows, colWidths=[30 * mm, 27 * mm, 50 * mm] * 2)
    paired = []
    for i in range(0, len(card_rows), 2):
        paired.append(card_rows[i] + (card_rows[i + 1] if i + 1 < len(card_rows) else ["", "", ""]))
    card_table = Table(paired, colWidths=[28 * mm, 25 * mm, 42 * mm, 28 * mm, 25 * mm, 42 * mm])
    card_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F9FC")),
        ("BOX", (0, 0), (-1, -1), 0.35, colors.HexColor("#B7C9D6")),
        ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D9EAF7")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TEXTCOLOR", (1, 0), (1, -1), colors.HexColor("#17365D")),
        ("TEXTCOLOR", (4, 0), (4, -1), colors.HexColor("#17365D")),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
    ]))
    story += [card_table, Spacer(1, 4 * mm)]

    story += [Paragraph("Personnel Statistics", section)]
    story.append(two_tables(
        "Employees by Department",
        [["Department", "Active", "Inactive", "Total"]] + [[r.get("label"), whole(r.get("active")), whole(r.get("inactive")), whole(r.get("total"))] for r in charts.get("byDepartment", [])[:10]],
        "Employees by Age Group",
        [["Age Group", "Employees"]] + [[r.get("label"), whole(r.get("total"))] for r in charts.get("byAgeGroup", [])],
        small,
    ))
    story += [Spacer(1, 3 * mm)]
    story.append(two_tables(
        "Employment Status",
        [["Status", "Active", "Inactive", "Total"]] + [[r.get("label"), whole(r.get("active")), whole(r.get("inactive")), whole(r.get("total"))] for r in charts.get("byEmploymentStatus", [])],
        "Top Positions",
        [["Position", "Employees"]] + [[Paragraph(txt(r.get("label")), small), whole(r.get("total"))] for r in charts.get("topPositions", [])[:10]],
        small,
    ))

    story += [Spacer(1, 4 * mm), Paragraph("Plantilla Analytics", section)]
    story.append(two_tables(
        "Plantilla by Division",
        [["Division", "Occupied", "Vacant", "Active"]] + [[Paragraph(txt(r.get("label")), small), whole(r.get("occupied")), whole(r.get("vacant")), whole(r.get("active"))] for r in charts.get("plantillaByDivision", [])[:10]],
        "Plantilla by Salary Grade",
        [["SG", "Occupied", "Vacant", "Active"]] + [[r.get("label"), whole(r.get("occupied")), whole(r.get("vacant")), whole(r.get("active"))] for r in charts.get("plantillaBySalaryGrade", [])[:14]],
        small,
    ))

    story += [PageBreak(), Paragraph("Plantilla Item Listing", section)]
    item_rows = [["Item No.", "Position", "SG", "Division / Section", "Status", "Occupancy", "Occupant"]]
    for item in (tables.get("plantillaItems") or [])[:120]:
        item_rows.append([
            txt(item.get("itemNumber")),
            Paragraph(txt(item.get("positionTitle")), small),
            txt(item.get("salaryGrade")),
            Paragraph(txt(item.get("division")) + (" / " + txt(item.get("section")) if item.get("section") else ""), small),
            txt(item.get("itemStatus")),
            txt(item.get("occupancyStatus")),
            Paragraph(txt(item.get("occupantName"), "-"), small),
        ])
    item_table = Table(item_rows, repeatRows=1, colWidths=[25 * mm, 58 * mm, 12 * mm, 54 * mm, 23 * mm, 24 * mm, 55 * mm])
    item_table.setStyle(base_table_style())
    story.append(item_table)

    def footer(canvas, doc):
        canvas.saveState()
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(colors.grey)
        canvas.drawString(10 * mm, 5 * mm, "Generated through STRH HRIS")
        canvas.drawRightString(287 * mm, 5 * mm, f"Page {doc.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)


def base_table_style():
    return TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#17365D")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONT", (0, 0), (-1, 0), "Helvetica-Bold", 6.7),
        ("FONT", (0, 1), (-1, -1), "Helvetica", 6.5),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B7C9D6")),
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F9FC")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ])


def table_block(title, rows, small):
    title_row = [[Paragraph(f"<b>{title}</b>", small)]]
    tbl = Table(title_row + rows, repeatRows=2)
    tbl.setStyle(base_table_style())
    tbl.setStyle(TableStyle([
        ("SPAN", (0, 0), (-1, 0)),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#D9EAF7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.HexColor("#17365D")),
    ]))
    return tbl


def two_tables(left_title, left_rows, right_title, right_rows, small):
    left = table_block(left_title, left_rows, small)
    right = table_block(right_title, right_rows, small)
    outer = Table([[left, right]], colWidths=[136 * mm, 136 * mm])
    outer.setStyle(TableStyle([("VALIGN", (0, 0), (-1, -1), "TOP")]))
    return outer


def main():
    if len(sys.argv) != 4:
        raise SystemExit("usage: personnel_plantilla_report.py input.json output format")
    with open(sys.argv[1], encoding="utf-8-sig") as f:
        data = json.load(f)
    os.makedirs(os.path.dirname(sys.argv[2]), exist_ok=True)
    if sys.argv[3] == "xlsx":
        xlsx(data, sys.argv[2])
    elif sys.argv[3] == "pdf":
        pdf(data, sys.argv[2])
    else:
        raise SystemExit("format must be xlsx or pdf")


if __name__ == "__main__":
    main()
