import json
import sys
from datetime import datetime, date, time

from openpyxl import load_workbook


def text(value):
    if value is None:
        return ""
    return str(value).strip()


def date_text(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    raw = text(value)
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%m-%d-%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(raw, fmt).date().isoformat()
        except ValueError:
            pass
    return raw[:10] if len(raw) >= 10 else ""


def time_text(value):
    if isinstance(value, datetime):
        return value.time().strftime("%H:%M:%S")
    if isinstance(value, time):
        return value.strftime("%H:%M:%S")
    raw = text(value)
    for fmt in ("%H:%M:%S", "%H:%M", "%I:%M:%S %p", "%I:%M %p"):
        try:
            return datetime.strptime(raw, fmt).time().strftime("%H:%M:%S")
        except ValueError:
            pass
    return raw


def main():
    if len(sys.argv) != 2:
        raise SystemExit("Usage: dtr_parse.py <workbook.xlsx>")

    workbook = load_workbook(sys.argv[1], data_only=True, read_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        print("[]")
        return

    header = [text(value).lower().replace(" ", "").replace("_", "") for value in rows[0]]
    has_header = any(key in header for key in ("employee", "employeeno", "employeeid", "date", "datetime", "time"))
    data_rows = rows[1:] if has_header else rows
    punches = []

    def get(row, *names):
        if not has_header:
            return ""
        for name in names:
            if name in header:
                index = header.index(name)
                return row[index] if index < len(row) else ""
        return ""

    for row in data_rows:
        if not any(text(value) for value in row):
            continue
        employee_no = text(get(row, "employeeno", "employeeid", "employee", "id")) if has_header else text(row[0] if len(row) > 0 else "")
        direct = get(row, "datetime", "punchat", "createdat") if has_header else ""
        if employee_no and direct:
            punch_at = f"{date_text(direct)} {time_text(direct)}" if isinstance(direct, datetime) else text(direct)
            punches.append({"employeeNo": employee_no, "punchAt": punch_at, "raw": [text(value) for value in row]})
            continue

        work_date = date_text(get(row, "date", "workdate")) if has_header else date_text(row[1] if len(row) > 1 else "")
        time_values = []
        if has_header:
            for key in ("amin", "amout", "pmin", "pmout", "time"):
                value = get(row, key)
                if text(value):
                    time_values.append(value)
        else:
            for value in row[2:6]:
                if text(value):
                    time_values.append(value)

        for value in time_values:
            punch_time = time_text(value)
            if employee_no and work_date and punch_time:
                punches.append({"employeeNo": employee_no, "punchAt": f"{work_date} {punch_time}", "raw": [text(item) for item in row]})

    print(json.dumps(punches))


if __name__ == "__main__":
    main()
