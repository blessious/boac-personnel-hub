import json, os, sys
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import landscape, A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

def txt(v): return "" if v is None else str(v)
def money(v):
    try: return float(v) if v not in (None,"") else None
    except: return None

def employee_name(e):
    parts=[e.get("firstname"),e.get("middlename"),e.get("lastname"),e.get("name_ext")]
    return " ".join(txt(x).strip() for x in parts if txt(x).strip())

def xlsx(data,out):
    wb=Workbook(); ws=wb.active; ws.title="Service Record"; ws.sheet_view.showGridLines=False
    navy="17365D"; blue="D9EAF7"; gray="E7E6E6"; red="9C0006"; thin=Side(style="thin",color="A6A6A6")
    ws.merge_cells("A1:L1"); ws["A1"]=txt(data.get("agency",{}).get("name") or "HR Personnel Management System"); ws["A1"].font=Font(name="Arial",size=15,bold=True,color="FFFFFF"); ws["A1"].fill=PatternFill("solid",fgColor=navy); ws["A1"].alignment=Alignment(horizontal="center"); ws.row_dimensions[1].height=26
    ws.merge_cells("A2:L2"); ws["A2"]="GENERIC SERVICE RECORD"; ws["A2"].font=Font(name="Arial",size=13,bold=True,color=navy); ws["A2"].alignment=Alignment(horizontal="center")
    ws.merge_cells("A3:L3"); ws["A3"]=data.get("notice",""); ws["A3"].font=Font(name="Arial",size=9,bold=True,color=red); ws["A3"].alignment=Alignment(horizontal="center")
    e=data.get("employee",{}); info=[("Employee",employee_name(e)),("Employee No.",e.get("employee_no")),("Current Department",e.get("department")),("Generated",data.get("generatedAt","")[:19].replace("T"," "))]
    for idx,(label,value) in enumerate(info):
        row=5+idx//2; col=1+(idx%2)*6; ws.cell(row,col,label).font=Font(bold=True,color=navy); ws.merge_cells(start_row=row,start_column=col+1,end_row=row,end_column=col+5); ws.cell(row,col+1,txt(value))
    headers=["From","To","Position / Designation","Department / Office","Status / Action","Annual Salary","SG","Step","Item No.","Branch","LWOP","Source"]
    header_row=8
    for c,h in enumerate(headers,1):
        cell=ws.cell(header_row,c,h); cell.fill=PatternFill("solid",fgColor=navy); cell.font=Font(name="Arial",size=9,bold=True,color="FFFFFF"); cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=Border(bottom=thin)
    for r_idx,r in enumerate(data.get("records",[]),header_row+1):
        vals=[r.get("serviceFrom"),r.get("serviceTo") or "Present",r.get("positionTitle"),r.get("department"),r.get("appointmentStatus"),money(r.get("annualSalary")),r.get("salaryGrade"),r.get("salaryStep"),r.get("itemNumber"),r.get("branch"),r.get("leaveWithoutPay"),r.get("source")]
        for c,v in enumerate(vals,1):
            cell=ws.cell(r_idx,c,v); cell.font=Font(name="Arial",size=9); cell.alignment=Alignment(vertical="top",wrap_text=True); cell.border=Border(bottom=Side(style="hair",color="D9D9D9"))
            if c==6 and v is not None: cell.number_format='₱#,##0.00'
        if r_idx%2==0:
            for c in range(1,13): ws.cell(r_idx,c).fill=PatternFill("solid",fgColor="F5F9FC")
    widths=[12,12,25,25,18,15,7,7,14,12,12,12]
    for c,w in enumerate(widths,1): ws.column_dimensions[get_column_letter(c)].width=w
    ws.freeze_panes="A9"; ws.auto_filter.ref=f"A8:L{max(8,8+len(data.get('records',[])))}"; ws.print_title_rows="1:8"; ws.page_setup.orientation="landscape"; ws.page_setup.fitToWidth=1; ws.sheet_properties.pageSetUpPr.fitToPage=True
    note_row=10+len(data.get("records",[])); ws.merge_cells(start_row=note_row,start_column=1,end_row=note_row,end_column=12); ws.cell(note_row,1,"This generic export is for data validation. Use the STRH-approved template when supplied.").font=Font(italic=True,size=8,color="666666")
    wb.save(out)

def pdf(data,out):
    doc=SimpleDocTemplate(out,pagesize=landscape(A4),leftMargin=10*mm,rightMargin=10*mm,topMargin=9*mm,bottomMargin=10*mm,title="Generic Service Record")
    styles=getSampleStyleSheet(); title=ParagraphStyle("title",parent=styles["Title"],fontName="Helvetica-Bold",fontSize=13,leading=15,alignment=TA_CENTER,textColor=colors.HexColor("#17365D")); small=ParagraphStyle("small",parent=styles["BodyText"],fontSize=6.5,leading=8); notice=ParagraphStyle("notice",parent=small,alignment=TA_CENTER,textColor=colors.HexColor("#9C0006"),fontName="Helvetica-Bold")
    e=data.get("employee",{}); story=[Paragraph(txt(data.get("agency",{}).get("name") or "HR Personnel Management System"),title),Paragraph("GENERIC SERVICE RECORD",title),Paragraph(data.get("notice",""),notice),Spacer(1,4*mm)]
    info=[["Employee",employee_name(e),"Employee No.",txt(e.get("employee_no"))],["Department",txt(e.get("department")),"Generated",txt(data.get("generatedAt",""))[:19].replace("T"," ")]]
    it=Table(info,colWidths=[22*mm,75*mm,25*mm,55*mm]);it.setStyle(TableStyle([("FONT",(0,0),(-1,-1),"Helvetica",8),("FONT",(0,0),(0,-1),"Helvetica-Bold",8),("FONT",(2,0),(2,-1),"Helvetica-Bold",8),("BOTTOMPADDING",(0,0),(-1,-1),3)]));story+=[it,Spacer(1,3*mm)]
    headers=["From","To","Position / Designation","Department / Office","Status / Action","Annual Salary","SG/Step","Item No.","LWOP","Source"]
    rows=[headers]
    for r in data.get("records",[]):
        sal=money(r.get("annualSalary")); rows.append([txt(r.get("serviceFrom")),txt(r.get("serviceTo") or "Present"),Paragraph(txt(r.get("positionTitle")),small),Paragraph(txt(r.get("department")),small),Paragraph(txt(r.get("appointmentStatus")),small),f"{sal:,.2f}" if sal is not None else "",f"{txt(r.get('salaryGrade'))}/{txt(r.get('salaryStep'))}".strip("/"),txt(r.get("itemNumber")),txt(r.get("leaveWithoutPay")),txt(r.get("source"))])
    table=Table(rows,repeatRows=1,colWidths=[18*mm,18*mm,43*mm,43*mm,31*mm,25*mm,14*mm,23*mm,18*mm,18*mm])
    table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),colors.HexColor("#17365D")),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONT",(0,0),(-1,0),"Helvetica-Bold",6.5),("FONT",(0,1),(-1,-1),"Helvetica",6.5),("GRID",(0,0),(-1,-1),0.25,colors.HexColor("#B7C9D6")),("VALIGN",(0,0),(-1,-1),"TOP"),("ALIGN",(0,0),(1,-1),"CENTER"),("ALIGN",(5,1),(7,-1),"RIGHT"),("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F5F9FC")]),("TOPPADDING",(0,0),(-1,-1),3),("BOTTOMPADDING",(0,0),(-1,-1),3)]));story.append(table);story+=[Spacer(1,3*mm),Paragraph("This generic export is for data validation. Use the STRH-approved template when supplied.",notice)]
    def footer(canvas,doc): canvas.saveState();canvas.setFont("Helvetica",7);canvas.setFillColor(colors.grey);canvas.drawRightString(287*mm,5*mm,f"Page {doc.page}");canvas.restoreState()
    doc.build(story,onFirstPage=footer,onLaterPages=footer)

def main():
    if len(sys.argv)!=4: raise SystemExit("usage: service_record_export.py input.json output format")
    with open(sys.argv[1],encoding="utf-8") as f:data=json.load(f)
    os.makedirs(os.path.dirname(sys.argv[2]),exist_ok=True)
    (xlsx if sys.argv[3]=="xlsx" else pdf)(data,sys.argv[2])
if __name__=="__main__":main()
